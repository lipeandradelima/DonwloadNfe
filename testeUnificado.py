# -*- coding: utf-8 -*-
"""
Programa com interface gráfica para baixar XMLs de NF-e da SEFAZ.

Versão Final e Otimizada v7:
- CORREÇÃO: Botão 'Abrir Pasta' agora funciona de forma confiável em
  diferentes sistemas operacionais (Windows, macOS, Linux).
- NOVO: Permite ao usuário alterar a pasta de destino dos XMLs.
- NOVO: Indica a pasta de destino dos XMLs e adiciona um botão para abri-la.
- NOVO: Permite abrir o XML baixado com um duplo clique na linha da tabela.
- MELHORIA: Status unificado para "Baixado" após a confirmação do download.
- Multithreading: Interface gráfica 100% responsiva durante o download.

Autor: Contribuição do Usuário / Gemini (Parceiro de Programacao)
Instalação: pip install requests lxml certifi openpyxl cryptography
"""

import os
import time
import gzip
import base64
import logging
import re
import tempfile
import threading
import queue
import webbrowser
import sys
import subprocess
from typing import Tuple, Optional, Dict

# Módulos de terceiros
import requests
import certifi
from lxml import etree
import openpyxl
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.serialization import pkcs12

# Módulos da interface gráfica
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk

# --- CONFIGURAÇÕES GLOBAIS ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)


# --- NÚCLEO DE LÓGICA (SEPARADO DA GUI) ---

def extrair_info_certificado(caminho_pfx: str, senha: str, silent=False) -> Optional[Dict[str, any]]:
    """Carrega um .pfx, extrai CNPJ, chave e certificado. Retorna None em caso de falha."""
    try:
        with open(caminho_pfx, "rb") as f:
            pfx_data = f.read()
        senha_bytes = senha.encode('utf-8')
        private_key, certificate, _ = pkcs12.load_key_and_certificates(pfx_data, senha_bytes)

        subject_str = str(certificate.subject)
        match = re.search(r'CN=.*?:(\d{14})', subject_str)
        cnpj = match.group(1) if match else None

        key_pem = private_key.private_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PrivateFormat.PKCS8,
            encryption_algorithm=serialization.NoEncryption()
        )
        cert_pem = certificate.public_bytes(encoding=serialization.Encoding.PEM)

        return {"cnpj": cnpj, "key_pem": key_pem, "cert_pem": cert_pem}

    except ValueError:
        if not silent: messagebox.showerror("Erro de Certificado", "A senha do certificado está incorreta.")
        return None
    except Exception as e:
        if not silent: messagebox.showerror("Erro de Certificado",
                                            f"Não foi possível processar o arquivo PFX.\n\nDetalhe: {e}")
        return None


def baixar_xml_por_chave(chave: str, cnpj: str, cert_files: Tuple[str, str], config: Dict, saida_dir: str) -> Dict[
    str, any]:
    """Executa a lógica de download para uma única chave de acesso, usando as configurações passadas."""
    try:
        ns_nfe = "http://www.portalfiscal.inf.br/nfe"
        ns_wsdl = "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe"
        ns_soap = "http://www.w3.org/2003/05/soap-envelope"

        distDFeInt = etree.Element("distDFeInt", versao="1.01", xmlns=ns_nfe)
        etree.SubElement(distDFeInt, "tpAmb").text = config['TP_AMB']
        etree.SubElement(distDFeInt, "cUFAutor").text = config['UF_CODIGO_IBGE']
        etree.SubElement(distDFeInt, "CNPJ").text = cnpj
        consChNFe = etree.SubElement(distDFeInt, "consChNFe")
        etree.SubElement(consChNFe, "chNFe").text = chave

        nfeDistDFeInteresse = etree.Element("nfeDistDFeInteresse", xmlns=ns_wsdl)
        nfeDadosMsg = etree.SubElement(nfeDistDFeInteresse, "nfeDadosMsg")
        nfeDadosMsg.append(distDFeInt)

        soap_envelope = etree.Element(f"{{{ns_soap}}}Envelope")
        soap_body = etree.SubElement(soap_envelope, f"{{{ns_soap}}}Body")
        soap_body.append(nfeDistDFeInteresse)

        payload = etree.tostring(soap_envelope)
        headers = {'Content-Type': 'application/soap+xml; charset=utf-8'}
        cert_path, key_path = cert_files

        response = requests.post(config['URL_WEBSERVICE'], data=payload, headers=headers, cert=(cert_path, key_path),
                                 timeout=config['TIMEOUT'])
        response.raise_for_status()

        root = etree.fromstring(response.content)
        namespaces = {'nfe': ns_nfe}
        retorno_element = root.find('.//nfe:retDistDFeInt', namespaces=namespaces)

        if retorno_element is None:
            return {"status": "Erro", "motivo": "Resposta da SEFAZ em formato inesperado."}

        cStat = retorno_element.find('nfe:cStat', namespaces=namespaces).text
        xMotivo = retorno_element.find('nfe:xMotivo', namespaces=namespaces).text

        if cStat == '138':
            docZip_element = retorno_element.find('.//nfe:docZip', namespaces=namespaces)
            if docZip_element is not None:
                xml_compactado_b64 = docZip_element.text
                xml_compactado = base64.b64decode(xml_compactado_b64)
                xml_conteudo = gzip.decompress(xml_compactado)

                nome_arquivo = f"{chave}-nfe.xml"
                caminho_arquivo = os.path.join(saida_dir, nome_arquivo)
                with open(caminho_arquivo, "wb") as f:
                    f.write(xml_conteudo)
                return {"status": "Baixado", "motivo": f"Salvo em: {saida_dir}"}
            else:
                return {"status": "Erro", "motivo": "cStat 138, mas o XML não foi retornado."}
        else:
            return {"status": "Erro", "motivo": f"({cStat}) {xMotivo}"}

    except requests.exceptions.HTTPError as e:
        return {"status": "Erro Fatal", "motivo": f"Erro de comunicação: {e.response.status_code}"}
    except Exception as e:
        return {"status": "Erro Fatal", "motivo": str(e)}


# --------------------------- CLASSE DA APLICAÇÃO (INTERFACE GRÁFICA) ---------------------------
class App(tk.Tk):
    """Classe principal da aplicação com interface gráfica."""
    UF_CODIGO_IBGE = "41"
    TP_AMB = "1"
    URL_WEBSERVICE = 'https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx'
    TIMEOUT = 60
    SLEEP_ENTRE_REQUISICOES = 1.0

    def __init__(self):
        super().__init__()
        self.title("Download de NF-e por Chave de Acesso")
        self.geometry("950x650")

        self.saida_dir = os.path.join(BASE_DIR, "XMLs_Baixados")
        os.makedirs(self.saida_dir, exist_ok=True)

        self.chaves_a_processar = []
        self.key_temp_path = None
        self.cert_temp_path = None
        self.download_em_andamento = False

        self.update_queue = queue.Queue()
        self.stop_signal = threading.Event()

        self.criar_widgets()

    def criar_widgets(self):
        """Constrói todos os elementos da interface gráfica."""
        style = ttk.Style(self)
        bg_color, fg_color, entry_bg = '#2E2E2E', '#FFFFFF', '#3C3C3C'
        self.configure(bg=bg_color)
        style.theme_use('clam')
        style.configure('.', background=bg_color, foreground=fg_color, fieldbackground=entry_bg, borderwidth=1)
        style.map('.', background=[('active', '#555555')])
        style.configure('TFrame', background=bg_color)
        style.configure('TLabel', background=bg_color, foreground=fg_color)
        style.configure('TLabelFrame.Label', background=bg_color, foreground=fg_color)
        style.configure('TButton', background='#0078D7', foreground=fg_color)
        style.map('TButton', background=[('active', '#005A9E')])
        style.configure("Treeview", background=entry_bg, foreground=fg_color, rowheight=25, fieldbackground=entry_bg)
        style.map('Treeview', background=[('selected', '#005A9E')])
        style.configure("Treeview.Heading", background="#555555", foreground=fg_color, relief="flat")

        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(expand=True, fill="both")

        info_frame = ttk.LabelFrame(main_frame, text="Informações da Empresa", padding="10")
        info_frame.pack(fill="x", expand=False, pady=5)
        info_frame.grid_columnconfigure(1, weight=1)
        ttk.Label(info_frame, text="Certificado (.pfx):").grid(row=0, column=0, sticky="w", pady=2, padx=(0, 5))
        self.entry_pfx = ttk.Entry(info_frame, width=40)
        self.entry_pfx.grid(row=0, column=1, pady=2, sticky="ew")
        self.btn_selecionar_pfx = ttk.Button(info_frame, text="Selecionar...", command=self.selecionar_pfx)
        self.btn_selecionar_pfx.grid(row=0, column=2, padx=5)
        ttk.Label(info_frame, text="Senha do Certificado:").grid(row=1, column=0, sticky="w", pady=2, padx=(0, 5))
        self.entry_senha = ttk.Entry(info_frame, width=50, show="*")
        self.entry_senha.grid(row=1, column=1, pady=2, sticky="ew")
        self.btn_testar_senha = ttk.Button(info_frame, text="Testar", command=self._validar_e_preencher_cnpj)
        self.btn_testar_senha.grid(row=1, column=2, padx=5)
        ttk.Label(info_frame, text="CNPJ da Empresa:").grid(row=2, column=0, sticky="w", pady=2, padx=(0, 5))
        self.entry_cnpj = ttk.Entry(info_frame, width=50)
        self.entry_cnpj.grid(row=2, column=1, pady=2, columnspan=2, sticky="ew")

        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill='x', pady=10)
        self.btn_carregar_chaves = ttk.Button(action_frame, text="Carregar Chaves (Excel)...",
                                              command=self.carregar_chaves_do_arquivo)
        self.btn_carregar_chaves.pack(side='left', padx=(0, 10))
        self.btn_acao_download = ttk.Button(action_frame, text="Iniciar Download", command=self.iniciar_download,
                                            state='disabled')
        self.btn_acao_download.pack(side='left', padx=(0, 10))

        chaves_frame = ttk.LabelFrame(main_frame, text="Resultados do Download", padding="10")
        chaves_frame.pack(fill="both", expand=True, pady=5)
        self.colunas = {"Status": 100, "Chave NF-e": 350, "Detalhe": 380, "Ação": 40}
        self.tree_chaves = ttk.Treeview(chaves_frame, columns=list(self.colunas.keys()), show='headings')
        for col, width in self.colunas.items():
            anchor = 'center' if col in ["Status", "Ação"] else 'w'
            stretch = (col == "Detalhe")
            self.tree_chaves.heading(col, text=col)
            self.tree_chaves.column(col, width=width, anchor=anchor, stretch=stretch)
        self.tree_chaves.tag_configure('sucesso', background='#1E4620', foreground='#A6D7A8')
        self.tree_chaves.tag_configure('erro', background='#5C2121', foreground='#F5B7B1')
        self.tree_chaves.tag_configure('info', background='#1B3A57', foreground='#AED6F1')
        self.tree_chaves.tag_configure('aguardando', background='#4A4A4A', foreground='#FFFFFF')
        self.tree_chaves.bind("<Button-1>", self._ao_clicar_na_arvore)
        self.tree_chaves.bind("<Double-1>", self._ao_duplo_clique_na_arvore)
        scrollbar = ttk.Scrollbar(chaves_frame, orient="vertical", command=self.tree_chaves.yview)
        self.tree_chaves.configure(yscrollcommand=scrollbar.set)
        self.tree_chaves.pack(side='left', expand=True, fill='both')
        scrollbar.pack(side='right', fill='y')

        output_frame = ttk.LabelFrame(main_frame, text="Pasta de Destino", padding="10")
        output_frame.pack(fill="x", expand=False, pady=(10, 0))
        output_frame.grid_columnconfigure(0, weight=1)
        self.entry_saida_dir = ttk.Entry(output_frame)
        self.entry_saida_dir.insert(0, os.path.realpath(self.saida_dir))
        self.entry_saida_dir.config(state='readonly')
        self.entry_saida_dir.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        self.btn_abrir_pasta = ttk.Button(output_frame, text="Abrir Pasta...", command=self._abrir_pasta_saida)
        self.btn_abrir_pasta.grid(row=0, column=1, padx=(0, 5))
        self.btn_alterar_destino = ttk.Button(output_frame, text="Alterar Destino...",
                                              command=self._selecionar_pasta_saida)
        self.btn_alterar_destino.grid(row=0, column=2)

        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill="x", expand=False, pady=(10, 0))
        self.status_label = ttk.Label(bottom_frame, text="Pronto.")
        self.status_label.pack(fill='x', expand=True)
        self.progress_bar = ttk.Progressbar(bottom_frame, orient='horizontal', mode='determinate')

    def _selecionar_pasta_saida(self):
        """Abre um diálogo para o usuário selecionar uma nova pasta de destino."""
        novo_caminho = filedialog.askdirectory(title="Selecione a nova pasta para salvar os XMLs",
                                               initialdir=self.saida_dir)
        if novo_caminho:
            self.saida_dir = novo_caminho
            os.makedirs(self.saida_dir, exist_ok=True)
            self.entry_saida_dir.config(state='normal')
            self.entry_saida_dir.delete(0, tk.END)
            self.entry_saida_dir.insert(0, os.path.realpath(self.saida_dir))
            self.entry_saida_dir.config(state='readonly')
            self.status_label.config(text=f"Pasta de destino alterada. Recarregue o Excel para atualizar o status.")
            messagebox.showinfo("Pasta Alterada",
                                "A pasta de destino foi atualizada. É recomendado carregar o arquivo Excel novamente para verificar quais chaves já existem no novo local.")

    def carregar_chaves_do_arquivo(self):
        """Lê um arquivo .xlsx e carrega as chaves, verificando no destino ATUAL."""
        filepath = filedialog.askopenfilename(title="Selecione o arquivo Excel com as chaves",
                                              filetypes=[("Excel", "*.xlsx")])
        if not filepath: return
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet = workbook.active
            chaves_encontradas = []
            coluna_de_chaves_idx = -1
            for row in sheet.iter_rows(max_row=20, values_only=True):
                for col_idx, cell_value in enumerate(row):
                    if isinstance(cell_value, str) and len("".join(filter(str.isdigit, cell_value))) == 44:
                        coluna_de_chaves_idx = col_idx;
                        break
                if coluna_de_chaves_idx != -1: break
            if coluna_de_chaves_idx == -1:
                return messagebox.showerror("Erro", "Nenhuma coluna com chaves de 44 dígitos foi encontrada.")
            for row in sheet.iter_rows(min_row=1, values_only=True):
                valor_celula = row[coluna_de_chaves_idx]
                if isinstance(valor_celula, str):
                    chave_limpa = "".join(filter(str.isdigit, valor_celula))
                    if len(chave_limpa) == 44: chaves_encontradas.append(chave_limpa)

            self.tree_chaves.delete(*self.tree_chaves.get_children())
            self.chaves_a_processar.clear()

            for chave in chaves_encontradas:
                caminho_arquivo = os.path.join(self.saida_dir, f"{chave}-nfe.xml")
                if os.path.exists(caminho_arquivo):
                    self.tree_chaves.insert("", "end", values=("Baixado", chave, "-", ""), tags=('info',))
                else:
                    item_id = self.tree_chaves.insert("", "end", values=("Aguardando", chave, "-", "❌"),
                                                      tags=('aguardando',))
                    self.chaves_a_processar.append({'item_id': item_id, 'chave': chave})

            if self.chaves_a_processar:
                self.btn_acao_download.config(state='normal')
                self.status_label.config(text=f"{len(self.chaves_a_processar)} chaves prontas para download.")
            else:
                self.btn_acao_download.config(state='disabled')
                self.status_label.config(text="Nenhuma chave nova para baixar.")
        except Exception as e:
            messagebox.showerror("Erro ao Ler Arquivo", f"Não foi possível processar o arquivo Excel.\n\nDetalhe: {e}")

    def _ao_duplo_clique_na_arvore(self, event):
        """Abre o XML correspondente ao item clicado, buscando na pasta de destino ATUAL."""
        item_id = self.tree_chaves.identify_row(event.y)
        if not item_id: return
        item_selecionado = self.tree_chaves.item(item_id)
        status = item_selecionado['values'][0]
        chave = item_selecionado['values'][1]
        if status == "Baixado" or 'info' in item_selecionado['tags']:
            nome_arquivo = f"{chave}-nfe.xml"
            caminho_arquivo = os.path.join(self.saida_dir, nome_arquivo)
            if os.path.exists(caminho_arquivo):
                try:
                    webbrowser.open(caminho_arquivo)
                except Exception as e:
                    messagebox.showerror("Erro ao Abrir", f"Não foi possível abrir o arquivo.\n\nDetalhe: {e}")
            else:
                messagebox.showwarning("Arquivo não Encontrado",
                                       f"O arquivo {nome_arquivo} não foi encontrado na pasta de destino atual.")

    def _abrir_pasta_saida(self):
        """Abre a pasta de destino ATUAL no gerenciador de arquivos do sistema (multiplataforma)."""
        caminho_normalizado = os.path.realpath(self.saida_dir)
        try:
            if sys.platform == 'win32':
                os.startfile(caminho_normalizado)
            elif sys.platform == 'darwin':  # macOS
                subprocess.Popen(['open', caminho_normalizado])
            else:  # Linux
                subprocess.Popen(['xdg-open', caminho_normalizado])
        except Exception as e:
            messagebox.showerror("Erro",
                                 f"Não foi possível abrir a pasta.\nVerifique se o caminho é válido.\n\nDetalhe: {e}")

    def _processo_de_download(self, cnpj: str, cert_files: Tuple[str, str], config: Dict):
        """Executado em uma thread separada para não travar a GUI."""
        final_msg = "Download de chaves concluído."
        try:
            total = len(self.chaves_a_processar)
            for i, item in enumerate(self.chaves_a_processar):
                if self.stop_signal.is_set():
                    final_msg = "Download interrompido pelo usuário.";
                    break
                self.update_queue.put(('update_tree', item['item_id'], ("Baixando...", item['chave'], "-", ""), ()))
                self.update_queue.put(('status', f"Baixando chave {i + 1}/{total}: {item['chave']}"))
                resultado = baixar_xml_por_chave(item['chave'], cnpj, cert_files, config, self.saida_dir)
                status = resultado.get("status", "Erro")
                motivo = resultado.get("motivo", "Sem detalhes")
                tag = ('sucesso',) if status == "Baixado" else ('erro',)
                self.update_queue.put(('update_tree', item['item_id'], (status, item['chave'], motivo, ""), tag))
                self.update_queue.put(('progress',))
                time.sleep(self.SLEEP_ENTRE_REQUISICOES)
            self.update_queue.put(('status', final_msg))
        except Exception as e:
            self.update_queue.put(('status', f"Erro na thread: {e}"))
        finally:
            self.update_queue.put(('finished',))

    def selecionar_pfx(self):
        filepath = filedialog.askopenfilename(title="Selecione o Certificado Digital",
                                              filetypes=[("Certificados PFX", "*.pfx")])
        if filepath: self.entry_pfx.delete(0, tk.END); self.entry_pfx.insert(0,
                                                                             filepath); self._validar_e_preencher_cnpj()

    def _validar_e_preencher_cnpj(self):
        pfx_path, password = self.entry_pfx.get(), self.entry_senha.get()
        if not os.path.exists(pfx_path) or not password: return
        info = extrair_info_certificado(pfx_path, password, silent=False)
        if info and info.get("cnpj"):
            self.entry_cnpj.delete(0, tk.END); self.entry_cnpj.insert(0, info["cnpj"])
        elif info:
            messagebox.showwarning("Atenção", "Certificado lido, mas não foi possível extrair o CNPJ.")

    def set_controls_state(self, state: str):
        is_running = (state == 'disabled')
        for widget in [self.btn_carregar_chaves, self.btn_selecionar_pfx, self.entry_pfx, self.entry_senha,
                       self.entry_cnpj, self.btn_testar_senha, self.btn_alterar_destino]:
            widget.config(state=state)
        if not is_running and self.chaves_a_processar:
            self.btn_acao_download.config(state='normal')
        else:
            self.btn_acao_download.config(state='disabled')

    def _preparar_execucao(self) -> Optional[Tuple[str, Tuple[str, str]]]:
        pfx_path, password = self.entry_pfx.get(), self.entry_senha.get()
        cnpj = "".join(filter(str.isdigit, self.entry_cnpj.get()))
        if not all([os.path.exists(pfx_path), password, len(cnpj) == 14]):
            messagebox.showerror("Erro de Validação", "Preencha o Certificado, a Senha e o CNPJ (14 dígitos).");
            return None
        cert_info = extrair_info_certificado(pfx_path, password)
        if not cert_info: return None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.key', mode='wb') as key_file:
                self.key_temp_path = key_file.name; key_file.write(cert_info["key_pem"])
            with tempfile.NamedTemporaryFile(delete=False, suffix='.cer', mode='wb') as cert_file:
                self.cert_temp_path = cert_file.name; cert_file.write(cert_info["cert_pem"])
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao criar arquivos de certificado temporários: {e}"); return None
        return cnpj, (self.cert_temp_path, self.key_temp_path)

    def _finalizar_execucao(self):
        if self.key_temp_path and os.path.exists(self.key_temp_path): os.remove(self.key_temp_path)
        if self.cert_temp_path and os.path.exists(self.cert_temp_path): os.remove(self.cert_temp_path)
        self.key_temp_path = self.cert_temp_path = None
        self.download_em_andamento = False
        self.stop_signal.clear()
        self.set_controls_state('normal')
        self.btn_acao_download.config(text="Iniciar Download", command=self.iniciar_download)

    def _ao_clicar_na_arvore(self, event):
        region = self.tree_chaves.identify_region(event.x, event.y)
        if region != "cell": return
        col_id = self.tree_chaves.identify_column(event.x)
        col_heading = self.tree_chaves.heading(col_id, 'text')
        if col_heading == "Ação":
            item_id = self.tree_chaves.identify_row(event.y)
            if not item_id: return
            item_para_remover = next((item for item in self.chaves_a_processar if item['item_id'] == item_id), None)
            if item_para_remover:
                self.chaves_a_processar.remove(item_para_remover)
                self.tree_chaves.delete(item_id)
                if self.chaves_a_processar:
                    self.status_label.config(text=f"{len(self.chaves_a_processar)} chaves prontas para download.")
                else:
                    self.status_label.config(text="Nenhuma chave para baixar."); self.btn_acao_download.config(
                        state='disabled')

    def iniciar_download(self):
        preparacao = self._preparar_execucao()
        if not preparacao: return
        self.download_em_andamento = True
        self.set_controls_state('disabled')
        self.btn_acao_download.config(text="Parar Download", command=self.parar_download, state='normal')
        self.progress_bar.pack(fill='x', expand=True, pady=5)
        self.progress_bar['maximum'] = len(self.chaves_a_processar)
        self.progress_bar['value'] = 0
        thread = threading.Thread(target=self._processo_de_download, args=(*preparacao, self.get_config()), daemon=True)
        thread.start()
        self.monitorar_fila()

    def get_config(self) -> Dict:
        return {'UF_CODIGO_IBGE': self.UF_CODIGO_IBGE, 'TP_AMB': self.TP_AMB, 'URL_WEBSERVICE': self.URL_WEBSERVICE,
                'TIMEOUT': self.TIMEOUT}

    def parar_download(self):
        self.status_label.config(text="Sinal de parada enviado... Aguardando a chave atual.")
        self.stop_signal.set()
        self.btn_acao_download.config(state='disabled')

    def monitorar_fila(self):
        try:
            while not self.update_queue.empty():
                msg = self.update_queue.get_nowait()
                msg_type, *args = msg
                if msg_type == 'status':
                    self.status_label.config(text=args[0])
                elif msg_type == 'update_tree':
                    self.tree_chaves.item(args[0], values=args[1], tags=args[2])
                elif msg_type == 'progress':
                    self.progress_bar.step()
                elif msg_type == 'finished':
                    self.progress_bar.pack_forget(); self._finalizar_execucao(); return
        except queue.Empty:
            pass
        if self.download_em_andamento: self.after(100, self.monitorar_fila)


if __name__ == "__main__":
    app = App()
    app.mainloop()
